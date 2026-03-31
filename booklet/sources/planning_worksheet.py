from __future__ import annotations

from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook

from booklet.models import SongAssignment, WorksheetService


FIELD_MAP = {
    "Description": "description",
    "Preacher": "preacher",
    "Celebrant": "celebrant",
    "Special Theme": "special_theme",
    "Special Elements": "special_elements",
    "Special Volunteers": "special_volunteers",
    "Readings": "compact_readings",
    "Collect": "collect",
    "Proper Prefaces": "proper_preface",
}


class PlanningWorksheetSource:
    def __init__(self, workbook_path: Path):
        self.workbook_path = workbook_path

    def get_service(self, service_date: date) -> WorksheetService | None:
        if not self.workbook_path.exists():
            return None

        workbook = load_workbook(self.workbook_path, data_only=True)
        try:
            for sheet in workbook.worksheets:
                service = self._service_from_sheet(sheet, service_date)
                if service:
                    return service
        finally:
            workbook.close()
        return None

    def _service_from_sheet(self, sheet, service_date: date) -> WorksheetService | None:
        column = self._find_date_column(sheet, service_date)
        if column is None:
            return None

        values: dict[str, str | None] = {}
        songs: list[SongAssignment] = []

        for row_index, row in enumerate(sheet.iter_rows(min_row=2), start=2):
            if not row:
                continue
            label = row[0].value
            if label is None:
                continue
            label = str(label).strip()
            if column - 1 >= len(row):
                continue
            raw_value = row[column - 1].value
            value = _normalize_cell(raw_value)
            if label in FIELD_MAP:
                values[FIELD_MAP[label]] = value
            elif label.startswith("Song "):
                songs.append(
                    SongAssignment(
                        slot=label.lower().replace(" ", "_"),
                        title=value,
                    )
                )

        return WorksheetService(
            service_date=service_date.isoformat(),
            description=values.get("description"),
            preacher=values.get("preacher"),
            celebrant=values.get("celebrant"),
            special_theme=values.get("special_theme"),
            special_elements=values.get("special_elements"),
            special_volunteers=values.get("special_volunteers"),
            compact_readings=values.get("compact_readings"),
            collect=values.get("collect"),
            proper_preface=values.get("proper_preface"),
            songs=songs,
        )

    def _find_date_column(self, sheet, service_date: date) -> int | None:
        header_rows = sheet.iter_rows(min_row=1, max_row=1)
        header = next(header_rows, None)
        if header is None:
            return None
        for column, cell in enumerate(header[1:], start=2):
            raw_value = cell.value
            if isinstance(raw_value, datetime) and raw_value.date() == service_date:
                return column
        return None


def _normalize_cell(value) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None
